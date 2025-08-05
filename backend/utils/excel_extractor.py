import pandas as pd;
import openpyxl;
import os;
import io;
import re;
from typing import Dict, Any, Tuple, Set, Generator;

class ExcelExtractor:
    
    OBJECT_TYPE_CHART = "CHART_DATA";
    OBJECT_TYPE_TABLE = "TABLE";
    
    def __init__(self, file_bytes: bytes, file_name:str):
        self.file_bytes = file_bytes;
        self.file_name = file_name;
        file_object = io.BytesIO(self.file_bytes);
        self.workbook = openpyxl.load_workbook(file_object, data_only=True);

    
    def _get_df_slice_from_range(self, df: pd.DataFrame, range_str: str) -> pd.DataFrame:
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_str);

        start_row_idx = min_row - 1;
        end_row_idx = max_row;
        start_col_idx = min_col - 1;
        end_col_idx = max_col;

        if min_row == 1:
            table_with_header = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx].copy();
            header = table_with_header.iloc[0];
            data_df = table_with_header[1:];
            data_df.columns = header;

        else:
            data_df = df.iloc[start_row_idx:end_row_idx, start_col_idx:end_col_idx].copy();
            header_row_index = start_row_idx - 1;
            header = df.iloc[header_row_index, start_col_idx:end_col_idx];
            data_df.columns = header;

        return data_df.reset_index(drop=True);
    
    def transform_grouped_data(self, df: pd.DataFrame, group_by_col: str, value_cols: list) -> dict:
        nested_data = {};
        if group_by_col not in df.columns or not all(c in df.columns for c in value_cols):
            return df.to_dict(orient='records');
        for group_name, group_df in df.groupby(group_by_col):
            nested_data[group_name] = group_df[value_cols].to_dict(orient='records');
        return nested_data;
    
    def _extract_chart_data(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, sheet_df: pd.DataFrame) -> Generator[Tuple[str, Dict[str, Any]], None, None]:
        for chart in worksheet._charts:
            if not chart.series:
                continue;

            min_col, min_row, max_col, max_row = float('inf'), float('inf'), float('-inf'), float('-inf');
            
            has_valid_range = False;
            for s in chart.series:
                val_ref = getattr(s, 'val', None);
                if val_ref and val_ref.numRef:
                    val_match = re.search(r'\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)', str(val_ref.numRef.f));
                    if val_match:
                        mc, mr, xc, xr = openpyxl.utils.range_boundaries(f"{val_match.group(1)}{val_match.group(2)}:{val_match.group(3)}{val_match.group(4)}");
                        min_col, min_row = min(min_col, mc), min(min_row, mr);
                        max_col, max_row = max(max_col, xc), max(max_row, xr);
                        has_valid_range = True;

                cat_ref = getattr(s, 'cat', None);
                if cat_ref:
                    cat_formula = getattr(cat_ref.multiLvlStrRef, 'f', getattr(cat_ref.strRef, 'f', None));
                    if cat_formula:
                        cat_match = re.search(r'\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)', str(cat_formula));
                        if cat_match:
                            mc, mr, xc, xr = openpyxl.utils.range_boundaries(f"{cat_match.group(1)}{cat_match.group(2)}:{cat_match.group(3)}{cat_match.group(4)}");
                            min_col, min_row = min(min_col, mc), min(min_row, mr);
                            max_col, max_row = max(max_col, xc), max(max_row, xr);
                            has_valid_range = True;
            
            if not has_valid_range:
                continue;

            total_range_str = f"{openpyxl.utils.get_column_letter(min_col)}{min_row}:{openpyxl.utils.get_column_letter(max_col)}{max_row}";
            
            try:
                source_table_df = self._get_df_slice_from_range(sheet_df, total_range_str);
                
                if source_table_df.shape[1] > 2:
                    grouping_column = source_table_df.columns[1];
                    value_columns = [source_table_df.columns[0], source_table_df.columns[-1]];
                    nested_chart_data = self.transform_grouped_data(source_table_df, grouping_column, value_columns);
                else:
                    nested_chart_data = source_table_df.to_dict(orient='records');
                
                chart_title = chart.title.text.rich.p[0].text[0].value or "Chart Tanpa Nama";
                x_axis_label = getattr(chart.x_axis, 'title', "");
                y_axis_label = getattr(chart.y_axis, 'title', "");

                chart_object = {
                    "objectType": self.OBJECT_TYPE_CHART,
                    "objectName": chart_title,
                    "sourceRange": f"{worksheet.title}!{total_range_str}",
                    "metadata": {
                        "xAxisLabel": str(x_axis_label) if x_axis_label else "",
                        "yAxisLabel": str(y_axis_label) if y_axis_label else ""
                    },
                    "data": nested_chart_data
                }
                yield total_range_str, chart_object;

            except Exception as e:
                print(f"Gagal memproses data chart '{chart.title}' di rentang {total_range_str}. Error: {e}");
                continue;
            
    def _find_data_tables(self, sheet_df: pd.DataFrame, processed_ranges: Set[str], sheet_name: str) -> Generator[Dict[str, Any], None, None]:
        processed_mask = pd.DataFrame(False, index=sheet_df.index, columns=sheet_df.columns);
        for range_str in processed_ranges:
            min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(range_str);
            processed_mask.iloc[min_row-1:max_row, min_col-1:max_col] = True;

        data_map = sheet_df.notna() & ~processed_mask;
        visited_mask = pd.DataFrame(False, index=data_map.index, columns=data_map.columns);

        for r_idx, row in data_map.iterrows():
            for c_idx, has_data in enumerate(row):
                if has_data and not visited_mask.iloc[r_idx, c_idx]:
                    max_r = r_idx;
                    while max_r + 1 < len(data_map) and data_map.iloc[max_r + 1].any():
                        max_r += 1;
                    
                    max_c = c_idx;
                    while max_c + 1 < len(data_map.columns) and data_map.iloc[r_idx:max_r+1, max_c + 1].any():
                        max_c += 1;
                    table_df = sheet_df.iloc[r_idx:max_r+1, c_idx:max_c+1].dropna(how='all', axis=0).dropna(how='all', axis=1);
                    visited_mask.iloc[r_idx:max_r+1, c_idx:max_c+1] = True;

                    if not table_df.empty:
                        table_df.columns = table_df.iloc[0];
                        table_df = table_df[1:].reset_index(drop=True);
                        
                        start_cell = f"{openpyxl.utils.get_column_letter(c_idx + 1)}{r_idx + 1}";
                        end_cell = f"{openpyxl.utils.get_column_letter(c_idx + table_df.shape[1])}{r_idx + 1 + len(table_df)}";
                        table_range_str = f"{start_cell}:{end_cell}";

                        yield {
                            "objectType": self.OBJECT_TYPE_TABLE,
                            "objectName": f"Tabel di {table_range_str}",
                            "sourceRange": f"{sheet_name}!{table_range_str}",
                            "data": table_df.to_dict(orient='records')
                        };
                        
    def generate_json_from_excel(self) -> Dict[str, Any]:
        result = {
            "fileName": self.file_name,
            "sheets": []
        };

        try:
            file_object = io.BytesIO(self.file_bytes);
            with pd.ExcelFile(file_object) as excel_file:
                for i, sheet_name in enumerate(self.workbook.sheetnames):
                    worksheet = self.workbook[sheet_name];
                    sheet_df = excel_file.parse(sheet_name, header=None);
                    sheet_df = sheet_df.astype(str);
                    sheet_data = {
                        "sheetName": sheet_name,
                        "sheetIndex": i,
                        "contentObjects": []
                    };
                    processed_ranges = set();

                    for range_str, chart_object in self._extract_chart_data(worksheet, sheet_df):
                        sheet_data["contentObjects"].append(chart_object);
                        processed_ranges.add(range_str);
                        
                    for table_object in self._find_data_tables(sheet_df, processed_ranges, sheet_name):
                        if table_object["data"]:
                            sheet_data["contentObjects"].append(table_object);

                    if sheet_data["contentObjects"]:
                        result["sheets"].append(sheet_data);
        finally:
            self.workbook.close();
        return result;

#excel_extractor = ExcelExtractor();